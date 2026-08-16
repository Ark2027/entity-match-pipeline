from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


MATCH_PRIORITY_COLUMNS = (
    "status",
    "status_note",
    "crm_source_name",
    "source_name",
    "crm_business_name",
    "live_business_name",
    "match_score",
    "primary_origination_date",
    "primary_origination_amount",
    "candidate_application_id",
    "origination_amounts",
    "application_created_date",
    "match_key",
)

MATCH_VISIBLE_COLUMNS = {
    "status",
    "status_note",
    "crm_source_name",
    "source_name",
    "crm_business_name",
    "live_business_name",
    "match_score",
    "primary_origination_date",
    "primary_origination_amount",
    "candidate_application_id",
    "origination_amounts",
    "application_created_date",
}


# Columns matching any of these are dropped before the workbook is written.
# Hiding a column in Excel is not redaction: it survives the file, and one
# right-click reveals it. Anything that should not travel with a shared
# workbook has to be removed, not concealed. Nothing currently produced by the
# pipeline matches these, and this is here to keep it that way.
SENSITIVE_COLUMN_PATTERNS = ("email", "phone", "ssn", "address_line", "date_of_birth", "dob")


def drop_sensitive_columns(frame: pd.DataFrame) -> pd.DataFrame:
    """Remove any column whose name suggests it carries personal detail."""
    if frame.empty:
        return frame
    drop = [
        column
        for column in frame.columns
        if any(pattern in str(column).lower() for pattern in SENSITIVE_COLUMN_PATTERNS)
    ]
    return frame.drop(columns=drop) if drop else frame


def _sanitize_for_excel(frame: pd.DataFrame) -> pd.DataFrame:
    sanitized = drop_sensitive_columns(frame.copy())
    for column in sanitized.columns:
        series = sanitized[column]
        if pd.api.types.is_datetime64tz_dtype(series):
            sanitized[column] = series.dt.tz_localize(None)
    return sanitized


def _reorder_match_columns(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    front = [column for column in MATCH_PRIORITY_COLUMNS if column in frame.columns]
    remaining = [column for column in frame.columns if column not in front]
    return frame[front + remaining]


def _autosize_columns(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
            max_length = min(max((len(value) for value in values), default=0) + 2, 50)
            sheet.column_dimensions[column_cells[0].column_letter].width = max_length
        if sheet.title in {"auto_accept", "review", "possible_matches"} and sheet.max_row >= 1:
            headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
            for header, column_index in headers.items():
                letter = get_column_letter(column_index)
                if header not in MATCH_VISIBLE_COLUMNS:
                    sheet.column_dimensions[letter].hidden = True
            status_column = headers.get("status")
            if status_column:
                letter = get_column_letter(status_column)
                validation = DataValidation(
                    type="list",
                    formula1='"new,confirmed,rejected,billed"',
                    allow_blank=False,
                )
                sheet.add_data_validation(validation)
                validation.add(f"{letter}2:{letter}{max(sheet.max_row, 2)}")
            status_note_column = headers.get("status_note")
            if status_note_column:
                sheet.column_dimensions[get_column_letter(status_note_column)].width = 24
            match_key_column = headers.get("match_key")
            if match_key_column:
                sheet.column_dimensions[get_column_letter(match_key_column)].hidden = True
            sheet.freeze_panes = "A2"
    workbook.save(workbook_path)


def build_output_path(output_dir: Path) -> Path:
    return output_dir / "match_review_current.xlsx"


def write_report(
    workbook_path: Path,
    *,
    summary: dict[str, object],
    auto_accept: pd.DataFrame,
    review: pd.DataFrame,
    possible_matches: pd.DataFrame,
    review_overflow: pd.DataFrame,
    exceptions: pd.DataFrame,
    duplicates: pd.DataFrame,
    not_surfaced: pd.DataFrame,
) -> None:
    summary_frame = pd.DataFrame(
        [{"metric": key, "value": value} for key, value in summary.items()]
    )
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        _sanitize_for_excel(summary_frame).to_excel(writer, sheet_name="summary", index=False)
        _sanitize_for_excel(_reorder_match_columns(auto_accept)).to_excel(writer, sheet_name="auto_accept", index=False)
        _sanitize_for_excel(_reorder_match_columns(review)).to_excel(writer, sheet_name="review", index=False)
        _sanitize_for_excel(_reorder_match_columns(possible_matches)).to_excel(writer, sheet_name="possible_matches", index=False)
        if not review_overflow.empty:
            _sanitize_for_excel(_reorder_match_columns(review_overflow)).to_excel(writer, sheet_name="review_overflow", index=False)
        _sanitize_for_excel(exceptions).to_excel(writer, sheet_name="exceptions", index=False)
        _sanitize_for_excel(duplicates).to_excel(writer, sheet_name="duplicates", index=False)
        if not not_surfaced.empty:
            _sanitize_for_excel(not_surfaced).to_excel(writer, sheet_name="not_surfaced", index=False)
    _autosize_columns(workbook_path)
